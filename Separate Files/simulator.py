import openpyxl as xl
from subprocess import call
import os
import sys

def gen_param_file(pdict, fldr_string):
    params_header = "#ifndef __params__\n#define __params__\n\n"
    params_header += "#define DIM %s\n" % pdict['DIM']
    params_header += "#define TIME_STEP %s\n" % pdict['TIME_STEP']
    params_header += "#define SIM_TIME %s\n" % pdict['SIM_TIME']
    params_header += "#define TOL %s\n" % pdict['TOL']
    params_header += "#define STEPS %s\n" % pdict['STEPS']
    params_header += "#define L_MEAN %s\n" % pdict['L_MEAN']
    params_header += "#define L_STD %s\n" % pdict['L_STD']
    params_header += "#define MAXBOUND %s\n" % pdict['MAXBOUND']
    params_header += "#define PAD %s\n" % pdict['PAD']
    params_header += "#define SACBONDS %s\n" % pdict['SACBONDS']
    params_header += "#define IMPLEMENT_PBC %s\n" % pdict['IMPLEMENT_PBC']
    params_header += '#define FLDR_STRING %s\n' % fldr_string
    params_header += "#define CRACKED %s\n" % pdict['CRACKED']
    params_header += "#define RATE_DAMAGE %s\n" % pdict['RATE_DAMAGE']
    params_header += "\n"
    params_header += "#if CRACKED\n#define PROB_REMOVAL %s\n" % pdict['PROB_REMOVAL'] +\
        "#else\n#define PROB_REMOVAL 0.0\n#endif\n"
    params_header += "\n#define __constants__\n\n"
    params_header += "#define kB %s\n" % pdict['kB']
    params_header += "#define b_poly %s\n" % pdict['b_poly']
    params_header += "#define T %s\n" % pdict['T']
    params_header += "#define ae %s\n" % pdict['ae']
    params_header += "#define delxe %s\n" % pdict['delxe']
    params_header += "#define af %s\n" % pdict['af']
    params_header += "#define delxf %s\n" % pdict['delxf']
    params_header += "\n#endif\n"
    # TODO will write to file params.h,
    f = open('params.h', 'w')
    f.write(params_header)
    f.close()
    return True


class xlreader():

    def __init__(self, fname):
        self.__wb = xl.load_workbook(filename=fname)
        self.ws = self.__wb.active
        self.columns = self.ws.max_column
        self.rows = self.ws.max_row
        self.param_list = self.setup_experiments()
        print "Ready!"

    def setup_experiments(self):
        list_of_params = []
        alphabet = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
        for j in range(1, self.columns):
            col = alphabet[j]
            list_of_params.append(self.read_column(col))
        return list_of_params

    def read_column(self, column='B'):
        params_dict = {}
        for i in range(1, self.rows + 1):
            # if self.ws['A'+str(i)].value is None:
            #     print("%d rows read!" % (i-1))
            #     break
            # else:
            params_dict[
                self.ws['A'+str(i)].value] = self.ws[column+str(i)].value
        return params_dict

    def compile(self,n):
        # TODO: subprocess for run
        for i, param in enumerate(self.param_list):
          fldr_string = '"set%d"'%(n+i)
          fldr_name = "set%d"%(n+i)
          if not os.path.exists(fldr_name):
            os.makedirs(fldr_name)
          gen_param_file(param, fldr_string)
          call("./compile.sh " + str(n+i), shell=True)
        return None

def main(n):
    a = xlreader("param_list.xlsx")
    a.compile(n)

if __name__ == '__main__':
    main(int(sys.argv[1]))
